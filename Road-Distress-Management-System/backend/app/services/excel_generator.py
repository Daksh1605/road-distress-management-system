"""
Excel Report Generation Service.
Utilizes openpyxl to compile professional, stakeholder-ready Excel reports (.xlsx)
comprising Executive Summary, Distress Records, Maintenance Tasks, and Analytics charts.
"""

import os
import logging
from datetime import datetime
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

from app.models.video import UploadedVideo
from app.models.distress import RoadDistress
from app.models.maintenance import MaintenanceTask

logger = logging.getLogger(__name__)


def generate_video_excel_report(db: Session, video_id: int) -> str:
    """
    Generates a comprehensive Excel report (.xlsx) for a processed video session.
    
    Args:
        db (Session): Database session.
        video_id (int): ID of the processed video.
        
    Returns:
        str: Relative path of the generated Excel file.
    """
    # 1. Fetch data from DB
    video = db.query(UploadedVideo).filter(UploadedVideo.id == video_id).first()
    if not video:
        raise ValueError(f"Video with ID {video_id} not found.")

    distresses = db.query(RoadDistress).filter(RoadDistress.video_id == video_id).all()
    
    # Resolve directories
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
    generated_dir = os.path.join(base_dir, "reports", "excel")
    os.makedirs(generated_dir, exist_ok=True)
    
    # Generate timestamp and filename
    timestamp_str = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    report_filename = f"report_video_{video_id}_{timestamp_str}.xlsx"
    full_path = os.path.join(generated_dir, report_filename)
    relative_path = os.path.relpath(full_path, base_dir).replace("\\", "/")

    # Initialize openpyxl Workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Reusable styles
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=13, bold=True, color="1F4E78")
    bold_font = Font(name=font_family, size=11, bold=True, color="000000")
    regular_font = Font(name=font_family, size=11, color="000000")
    italic_font = Font(name=font_family, size=9, italic=True, color="595959")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Classic Navy
    accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Soft steel blue
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid") # Light gray
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom_border = Border(bottom=Side(border_style="medium", color="1F4E78"))
    double_bottom_border = Border(bottom=Side(border_style="double", color="1F4E78"), top=thin_border_side)

    # ----------------------------------------------------
    # SHEET 1: Executive Summary
    # ----------------------------------------------------
    ws1 = wb.create_sheet(title="Executive Summary")
    ws1.views.sheetView[0].showGridLines = True

    # Title
    ws1.cell(row=1, column=1, value="Road Distress Management System").font = italic_font
    ws1.cell(row=2, column=1, value="Executive Summary Report").font = title_font
    
    # Metadata Block
    meta_rows = [
        ("Report Identifier", f"EXCEL-REP-VIDEO-{video_id}-{timestamp_str}"),
        ("Generated Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Source Video File", video.filename),
        ("Source Video ID", video_id),
        ("Analysis Run Status", video.processing_status.upper())
    ]
    
    start_row = 4
    for idx, (label, val) in enumerate(meta_rows):
        lbl_cell = ws1.cell(row=start_row + idx, column=1, value=label)
        lbl_cell.font = bold_font
        lbl_cell.fill = zebra_fill
        lbl_cell.border = thin_border
        
        val_cell = ws1.cell(row=start_row + idx, column=2, value=val)
        val_cell.font = regular_font
        val_cell.border = thin_border
        if label == "Analysis Run Status":
            val_cell.font = Font(name=font_family, size=11, bold=True, color="2E7D32" if val == "COMPLETED" else "C62828")

    # Gather tally stats
    total_distresses = len(distresses)
    critical_count = 0
    high_count = 0
    medium_count = 0
    low_count = 0

    for d in distresses:
        sev = d.severity.lower()
        if sev == "critical":
            critical_count += 1
        elif sev == "high":
            high_count += 1
        elif sev == "medium":
            medium_count += 1
        elif sev == "low":
            low_count += 1

    # Severity Summary Table
    ws1.cell(row=11, column=1, value="Distress Severity Distribution Summary").font = section_font
    
    summary_headers = ["Severity Level", "Anomaly Count", "Urgency Status"]
    for col_idx, text in enumerate(summary_headers):
        cell = ws1.cell(row=13, column=col_idx + 1, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    summary_data = [
        ("Critical", critical_count, "Emergency Action Required (P1)"),
        ("High", high_count, "Scheduled Repair Priority (P2)"),
        ("Medium", medium_count, "Routine Preventive (P3)"),
        ("Low", low_count, "Monitor / Periodic Audit (P4)"),
    ]
    
    for row_offset, (sev, count, urgency) in enumerate(summary_data):
        r = 14 + row_offset
        c1 = ws1.cell(row=r, column=1, value=sev)
        c1.font = bold_font
        c1.border = thin_border
        c1.alignment = Alignment(horizontal="left")
        
        c2 = ws1.cell(row=r, column=2, value=count)
        c2.font = regular_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal="right")
        
        c3 = ws1.cell(row=r, column=3, value=urgency)
        c3.font = regular_font
        c3.border = thin_border
        c3.alignment = Alignment(horizontal="left")
        
        # Color coding
        if sev == "Critical":
            c1.font = Font(name=font_family, size=11, bold=True, color="C62828")
        elif sev == "High":
            c1.font = Font(name=font_family, size=11, bold=True, color="EF6C00")

    # Add Total row
    r_total = 18
    t1 = ws1.cell(row=r_total, column=1, value="Total Logged Anomalies")
    t1.font = bold_font
    t1.fill = accent_fill
    t1.border = double_bottom_border
    
    t2 = ws1.cell(row=r_total, column=2, value=f"=SUM(B14:B17)")
    t2.font = bold_font
    t2.fill = accent_fill
    t2.border = double_bottom_border
    t2.alignment = Alignment(horizontal="right")
    
    t3 = ws1.cell(row=r_total, column=3, value="")
    t3.border = double_bottom_border
    t3.fill = accent_fill

    # Auto-adjust column widths for sheet 1
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip title row and total formula row in calculations to avoid huge widths
            if cell.row in [2, 18] or cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # ----------------------------------------------------
    # SHEET 2: Distress Records
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Distress Records")
    ws2.views.sheetView[0].showGridLines = True
    
    rec_headers = [
        "Distress ID", "Distress Type", "Severity", "Confidence Score",
        "Latitude", "Longitude", "Video Timestamp (s)", "Frame Number", "Prescribed Recommendation"
    ]
    
    # Headers formatting
    for col_idx, text in enumerate(rec_headers):
        cell = ws2.cell(row=1, column=col_idx + 1, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws2.row_dimensions[1].height = 28
        
    for row_idx, d in enumerate(distresses):
        r = 2 + row_idx
        # Query recommendation
        task = db.query(MaintenanceTask).filter(MaintenanceTask.distress_id == d.id).first()
        rec_text = task.recommendation if task else "Routine monitoring."
        
        # Populate columns
        ws2.cell(row=r, column=1, value=d.id).alignment = Alignment(horizontal="center")
        ws2.cell(row=r, column=2, value=d.distress_type.capitalize()).alignment = Alignment(horizontal="left")
        ws2.cell(row=r, column=3, value=d.severity.capitalize()).alignment = Alignment(horizontal="center")
        
        conf_cell = ws2.cell(row=r, column=4, value=d.confidence_score)
        conf_cell.number_format = "0.0%"
        conf_cell.alignment = Alignment(horizontal="right")
        
        lat_cell = ws2.cell(row=r, column=5, value=d.latitude)
        lat_cell.number_format = "0.00000"
        lat_cell.alignment = Alignment(horizontal="right")
        
        lon_cell = ws2.cell(row=r, column=6, value=d.longitude)
        lon_cell.number_format = "0.00000"
        lon_cell.alignment = Alignment(horizontal="right")
        
        ts_cell = ws2.cell(row=r, column=7, value=d.video_timestamp)
        if d.video_timestamp is not None:
            ts_cell.number_format = "0.0"
        ts_cell.alignment = Alignment(horizontal="right")
        
        f_cell = ws2.cell(row=r, column=8, value=d.frame_number)
        f_cell.alignment = Alignment(horizontal="right")
        
        ws2.cell(row=r, column=9, value=rec_text).alignment = Alignment(horizontal="left")
        
        # Zebra striping & borders
        for c in range(1, len(rec_headers) + 1):
            cell = ws2.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
                
            # Color-code severity label
            if c == 3:
                sev_val = str(cell.value).lower()
                if sev_val == "critical":
                    cell.font = Font(name=font_family, size=11, bold=True, color="C62828")
                elif sev_val == "high":
                    cell.font = Font(name=font_family, size=11, bold=True, color="EF6C00")

    # Column dimensions for sheet 2
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 13)

    # ----------------------------------------------------
    # SHEET 3: Maintenance Tasks
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Maintenance Tasks")
    ws3.views.sheetView[0].showGridLines = True

    maint_headers = [
        "Task ID", "Distress ID", "Priority Rank", "Prescribed Maintenance Action", 
        "Scheduling Status", "Estimated Response Window", "Estimated Repair Cost (₹)"
    ]
    
    # Headers formatting
    for col_idx, text in enumerate(maint_headers):
        cell = ws3.cell(row=1, column=col_idx + 1, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws3.row_dimensions[1].height = 28

    tasks_populated = 0
    for row_idx, d in enumerate(distresses):
        task = db.query(MaintenanceTask).filter(MaintenanceTask.distress_id == d.id).first()
        if not task:
            continue
            
        r = 2 + tasks_populated
        tasks_populated += 1
        
        ws3.cell(row=r, column=1, value=task.id).alignment = Alignment(horizontal="center")
        ws3.cell(row=r, column=2, value=task.distress_id).alignment = Alignment(horizontal="center")
        ws3.cell(row=r, column=3, value=task.priority.upper()).alignment = Alignment(horizontal="center")
        ws3.cell(row=r, column=4, value=task.recommendation).alignment = Alignment(horizontal="left")
        ws3.cell(row=r, column=5, value=task.status.upper()).alignment = Alignment(horizontal="center")
        ws3.cell(row=r, column=6, value=task.estimated_response_time or "N/A").alignment = Alignment(horizontal="left")
        
        cost_cell = ws3.cell(row=r, column=7, value=task.estimated_cost or 0)
        cost_cell.number_format = "₹#,##0"
        cost_cell.alignment = Alignment(horizontal="right")

        # Zebra striping & borders
        for c in range(1, len(maint_headers) + 1):
            cell = ws3.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
                
            # Color-code priority rank
            if c == 3:
                p_val = str(cell.value).upper()
                if p_val == "P1":
                    cell.font = Font(name=font_family, size=11, bold=True, color="C62828")
                elif p_val == "P2":
                    cell.font = Font(name=font_family, size=11, bold=True, color="EF6C00")

    # Column dimensions for sheet 3
    for col in ws3.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 13)

    # ----------------------------------------------------
    # SHEET 4: Analytics
    # ----------------------------------------------------
    ws4 = wb.create_sheet(title="Analytics")
    ws4.views.sheetView[0].showGridLines = True
    
    # 1. Distress Type Distribution
    ws4.cell(row=1, column=1, value="Distress Type Distribution").font = section_font
    ws4.cell(row=2, column=1, value="Distress Type").font = bold_font
    ws4.cell(row=2, column=1).fill = accent_fill
    ws4.cell(row=2, column=2, value="Anomalies").font = bold_font
    ws4.cell(row=2, column=2).fill = accent_fill
    
    distress_types = {}
    for d in distresses:
        t = d.distress_type.capitalize()
        distress_types[t] = distress_types.get(t, 0) + 1
        
    for idx, (dtype, count) in enumerate(distress_types.items()):
        r = 3 + idx
        ws4.cell(row=r, column=1, value=dtype).font = regular_font
        ws4.cell(row=r, column=2, value=count).font = regular_font
        
    last_dtype_row = 2 + len(distress_types) if distress_types else 2
    
    # 2. Severity Distribution
    ws4.cell(row=1, column=4, value="Severity Distribution").font = section_font
    ws4.cell(row=2, column=4, value="Severity Level").font = bold_font
    ws4.cell(row=2, column=4).fill = accent_fill
    ws4.cell(row=2, column=5, value="Detections").font = bold_font
    ws4.cell(row=2, column=5).fill = accent_fill
    
    severity_data = [
        ("Critical", critical_count),
        ("High", high_count),
        ("Medium", medium_count),
        ("Low", low_count)
    ]
    
    for idx, (sev, count) in enumerate(severity_data):
        r = 3 + idx
        ws4.cell(row=r, column=4, value=sev).font = regular_font
        ws4.cell(row=r, column=5, value=count).font = regular_font

    # 3. Detection Count Per Video across system
    ws4.cell(row=1, column=8, value="Detection Count Per Video (Overall)").font = section_font
    ws4.cell(row=2, column=8, value="Video ID").font = bold_font
    ws4.cell(row=2, column=8).fill = accent_fill
    ws4.cell(row=2, column=9, value="Filename").font = bold_font
    ws4.cell(row=2, column=9).fill = accent_fill
    ws4.cell(row=2, column=10, value="Anomalies").font = bold_font
    ws4.cell(row=2, column=10).fill = accent_fill
    
    all_videos = db.query(UploadedVideo).all()
    for idx, v in enumerate(all_videos):
        r = 3 + idx
        v_count = db.query(RoadDistress).filter(RoadDistress.video_id == v.id).count()
        ws4.cell(row=r, column=8, value=v.id).alignment = Alignment(horizontal="center")
        ws4.cell(row=r, column=9, value=v.filename).alignment = Alignment(horizontal="left")
        ws4.cell(row=r, column=10, value=v_count).alignment = Alignment(horizontal="right")
        ws4.cell(row=r, column=8).font = regular_font
        ws4.cell(row=r, column=9).font = regular_font
        ws4.cell(row=r, column=10).font = regular_font

    # Border framing for tables in analytics sheet
    for col_ranges in [(1, 2, last_dtype_row), (4, 5, 6), (8, 10, 2 + len(all_videos))]:
        start_c, end_c, max_r = col_ranges
        for row in range(2, max_r + 1):
            for col in range(start_c, end_c + 1):
                ws4.cell(row=row, column=col).border = thin_border

    # Add Bar Chart for Severity Distribution
    if total_distresses > 0:
        sev_chart = BarChart()
        sev_chart.type = "col"
        sev_chart.title = "Anomalies by Severity Level"
        sev_chart.style = 10
        sev_chart.y_axis.title = "Detections count"
        sev_chart.x_axis.title = "Severity Level"
        sev_chart.legend = None
        
        data_ref = Reference(ws4, min_col=5, min_row=2, max_row=6)
        cats_ref = Reference(ws4, min_col=4, min_row=3, max_row=6)
        sev_chart.add_data(data_ref, titles_from_data=True)
        sev_chart.set_categories(cats_ref)
        
        ws4.add_chart(sev_chart, "D9")

    # Add Pie Chart for Distress Type Distribution
    if distress_types:
        pie = PieChart()
        pie.title = "Distress Category Distribution"
        
        data_ref = Reference(ws4, min_col=2, min_row=2, max_row=last_dtype_row)
        cats_ref = Reference(ws4, min_col=1, min_row=3, max_row=last_dtype_row)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        
        ws4.add_chart(pie, "A9")

    # Set column widths for Sheet 4
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 13
    ws4.column_dimensions["C"].width = 5
    ws4.column_dimensions["D"].width = 16
    ws4.column_dimensions["E"].width = 13
    ws4.column_dimensions["F"].width = 5
    ws4.column_dimensions["G"].width = 5
    ws4.column_dimensions["H"].width = 10
    ws4.column_dimensions["I"].width = 28
    ws4.column_dimensions["J"].width = 13

    # Save to file
    wb.save(full_path)
    logger.info(f"Generated professional Excel report saved at: {full_path}")
    return relative_path
